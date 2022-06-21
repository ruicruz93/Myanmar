# - * - coding: utf-8 - * -
from osgeo import ogr
import openpyxl, win32com.client, sys, re
from openpyxl.utils import column_index_from_string

#*****************************************************************************

def records_show(excel_sheet, col_letters_tuple, start_row=1, end_row=0):#end_row=0 means last_row
    if end_row == 0:
        end_row = excel_sheet.max_row
    for i in range(start_row, end_row+1): #Rows and columns index start at 1.
        print(*tuple(excel_sheet.cell(row=i, column=column_index_from_string(col)).value for col in col_letters_tuple), sep='---->')
#********************************************************************************

def binarySearch(target, iterable, i_start = 0, i_end = 0, first_iter = True): #Iterable must be sorted beforehand
    if first_iter == True:
        i_end = len(iterable) - 1
    i_middle = round((i_end + i_start)/2)
    if i_start > i_end: #Target not in iterable
        return None    
    elif target == iterable[i_middle]:
        return i_middle
    elif target > iterable[i_middle]:
        return binarySearch(target, iterable, i_middle + 1, i_end, False)
    else:
        return binarySearch(target, iterable, i_start, i_middle - 1, False)
#********************************************************************************

#For those villages geometrically within a Village Tract different from the one assigned in their attributes, change it to that one
def vil_correct():
    in_source = fr'{directory}\VTs_Villages_Wards.gpkg'
    in_driver = ogr.GetDriverByName('GPKG')
    in_dataSource = in_driver.Open(in_source, 1) #Open for editing the database
    #Create a temporary layer with the problematic villages
    misplaced_vils_layer = in_dataSource.ExecuteSQL(
    '''SELECT v.Vill_Pcode, vt.VT_PCODE, vt.VT
    FROM Villages AS v
    JOIN VTs as vt
    ON ST_Contains(vt.geom, v.geom)
    WHERE v.VT != vt.VT
    ORDER BY v.Vill_Pcode ASC;''')
    #nr_fields = misplaced_vils_layer.GetLayerDefn().GetFieldCount()
    #for feat in misplaced_vils_layer:
        #print(*tuple(feat.GetField(i) for i in range(nr_fields)), sep= '--->')
    villages_layer = in_dataSource.GetLayerByName('Villages')
    villages_layer.StartTransaction()
    #Filter the Villages layer based on the codes of the misplaced villages
    villages_layer.SetAttributeFilter(f'Vill_Pcode IN ({",".join(tuple(str(feat.GetField(0)) for feat in misplaced_vils_layer))})')
    misplaced_vils_layer.ResetReading() #Should be done after each loop of any layer
    #Iterate over the now filtered villages layer and change each village's VT_Pcode and VT attribute values based on the equivalent misplaced village layer's record
    for village in villages_layer:
        village_code = village.GetField('Vill_Pcode')
        misplaced_vils_layer.SetAttributeFilter(f"Vill_Pcode = '{village_code}'") #Mind the single quotes wrapped around the attribute value
        ms_village = misplaced_vils_layer.GetNextFeature()
        village.SetField('VT_Pcode', ms_village.GetField('VT_PCODE'))
        village.SetField('VT', ms_village.GetField('VT'))
        villages_layer.SetFeature(village) #Make changes permanent
        misplaced_vils_layer.ResetReading()
    villages_layer.ResetReading()
    villages_layer.CommitTransaction()
    in_dataSource.ReleaseResultSet(misplaced_vils_layer) #Destroy the SQL layer
    in_dataSource = None
    print('Finished correcting Villages')

#*************************************************************************

#Add respective Census data from the Excel file to the Village Tracts and Wards layers
def add_pop_values():
    excel_file = r"C:\Users\Eu\UC\Git\Tese\MIMU_09Sep2015_MMR_Census_Data.xlsx"
    #Sort each excel sheet by respective code
    excel = win32com.client.Dispatch('Excel.Application')
    wb = excel.Workbooks.Open(excel_file)
    for i in range(wb.Worksheets.Count):
        sheet = wb.Worksheets[i]
        print(f"Working on sheet nr. {i+1}: {sheet.Name}")
        #Clear Data validation. In case there are some columns with data validation requirements that forbid cell value editing.
        sheet.Columns.Validation.Delete()
        working_range = sheet.Range("A2", sheet.Range("A2").End(2).End(4)) #Make sure working range's first row's cells are all filled, otherwise selection might fail
        working_range.Font.Size = 16
        working_range.Font.Name = "Times New Roman"
        working_range.Sort(Key1=sheet.Range('F1'), Order1=1, Orientation=1) #Sort Ascending, by columns, respectively
    excel.Application.DisplayAlerts = False
    sheet.SaveAs(excel_file)
    excel.Application.Quit()    
    #Now open the Excel file with Openpyxl
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    sheets = wb.sheetnames
    #Open the Geopackage
    in_source = fr'{directory}\VTs_Villages_Wards.gpkg'
    in_driver = ogr.GetDriverByName('GPKG')
    in_dataSource = in_driver.Open(in_source, 1)
    for l, layer_name in enumerate(('VTs', 'Wards')):
        if layer_name == 'VTs':
            code_field = 'VT_PCODE'
        else:
            code_field = 'Ward_Pcode'
        layer = in_dataSource.GetLayerByName(layer_name)
        layer.StartTransaction()
        fields_dic = {}
        for col in ('I', 'J', 'K'): #Excel columns with Total, Male and Female population counts, respectively
            #Create the respective field in the layer:
            fields_dic[col] = wb[sheets[l]][f'{col}1'].value
            layer.CreateField(ogr.FieldDefn(fields_dic[col], ogr.OFTInteger))
        #Both Ward and VT codes can be found on the F column in both sheets
        excel_codes = tuple(i[0].value for i in wb[sheets[l]]['F2':f'F{wb[sheets[l]].max_row}'])
        #Fill every record of current layer with respective values
        for r in range(layer.GetFeatureCount()):
            print(f"Analyzing feature {r+1} of {layer.GetFeatureCount()}")
            feature = layer.GetNextFeature()
            code = feature.GetField(code_field)
            #Find the sheet's row that matches the current feature
            row_nr = binarySearch(code, excel_codes) + 1 + 1 #+1 because excel rows start at 1, and other +1 to account for the sheet header
            for col in ('I', 'J', 'K'):
                feature.SetField(fields_dic[col], wb[sheets[l]][f'{col}{row_nr}'].value)
            layer.SetFeature(feature)
            feature = None
        layer.ResetReading()
        layer.CommitTransaction()    
    wb.close()
    in_dataSource = None
    print('Finished adding population values')
#*****************************************************************************

directory = re.compile(r'(.*)\\[^\\]+$').search(sys.argv[0]).group(1)

add_pop_values()
vil_correct()